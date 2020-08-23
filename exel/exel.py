import pandas as pd
import xlwt
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, PatternFill, NamedStyle, Fill

# workbook = load_workbook(filename="data/data_freeze.xlsx")
sheet = workbook.worksheets[0]

colunms = NamedStyle(name="Colunms")
colunms.alignment.horizontal = "center"
colunms.alignment.vertical = "center"
colunms.fill.start_color.index = Color.RED
colunms.fill.fill_type = Fill.FILL_SOLID
colunms.font = Font(name="Arial",bold=True)
colunms.border = Border(left=Side(style='thick'), 
                     right=Side(style='thick'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thick'))

rows = NamedStyle(name="Rows")
rows.font = Font(name="Arial",bold=False)
rows.border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
workbook.                     
# workbook.add_named_style(colunms)                    
# workbook.add_named_style(rows)

for x in sheet.iter_rows(min_row=2,max_row=len(sheet["A"])):
    for row in x:
        row.style = "Rows"
for y in sheet.iter_rows(max_row=1):
    for colunm in y:
        colunm.style = "Colunms"
        colunm.fill = PatternFill(start_color='00FFFF00',fill_type='solid')

workbook.save(filename="data/data_freeze.xlsx")
import pandas as pd
writer = pd.ExcelWriter("data/data1_excel.xlsx",engine="xlsxwriter")
wb = writer.book


# work with pandas
# pd.ExcelWriter("file")
# writer = pd.to_excel("file",engine="module")

